"""
Bolla PDF → Danea Excel Converter
Backend FastAPI — formato esatto da ENESCO_DANEA_FINALE.xlsx
"""

import json
import os
import re
import uuid
import tempfile
from pathlib import Path

import anthropic
import openpyxl
import pdfplumber
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from openpyxl.styles import Alignment, Font, PatternFill

# ─────────────────────────────────────────────
# App setup
# ─────────────────────────────────────────────
app = FastAPI(title="Bolla → Danea Converter")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

TEMP_DIR = Path(tempfile.gettempdir()) / "danea_converter"
TEMP_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
# Danea column structure (A–AX = 50 colonne)
# ─────────────────────────────────────────────
DANEA_HEADERS = [
    "Cod.", "Descrizione", "Tipologia", "Categoria", "Sottocategoria",
    "Cod. Udm", "Cod. Iva", "Listino 1 (ivato)", "Listino 2 (ivato)", "Listino 3 (ivato)",
    "Formula listino 1", "Formula listino 2", "Formula listino 3", "Note", "Cod. a barre",
    "Internet", "Produttore", "Descriz. web (Sorgente HTML)", "E-commerce", "Vendita Touch",
    "Extra 1", "Extra 2", "Extra 3", "Extra 4", "Cod. fornitore",
    "Fornitore", "Cod. prod. forn.", "Prezzo forn. (ivato)", "Note fornitura", "Ord. a multipli di",
    "Gg. ordine", "Scorta min.", "Ubicazione", "Tot. qtà caricata", "Tot. qtà scaricata",
    "Q.tà giacenza", "Q.tà impegnata", "Q.tà disponibile", "Q.tà in arrivo", "Vendita media mensile",
    "Stima data fine magazz.", "Stima data prossimo ordine", "Data primo carico", "Data ultimo carico",
    "Data ultimo scarico", "Costo medio d'acq.", "Ultimo costo d'acq.", "Prezzo medio vend.",
    "Stato magazzino", "Immagine",
]

# Larghezze colonne dal file di riferimento
COL_WIDTHS = {
    "A": 5.5,  "B": 11.66, "C": 9.5,  "D": 9.83,  "E": 14.16, "F": 10.0,
    "G": 8.5,  "H": 15.0,  "K": 16.16,"N": 5.5,   "O": 12.0,  "P": 8.16,
    "Q": 10.83,"R": 27.66, "S": 12.33, "T": 13.83, "U": 7.5,   "Y": 13.66,
    "Z": 9.5,  "AA": 15.16,"AB": 18.16,"AC": 13.66,"AD": 16.16,"AE": 10.5,
    "AF": 11.5,"AG": 11.0, "AH": 16.0, "AI": 17.0, "AJ": 13.16,"AK": 14.5,
    "AL": 15.16,"AM": 12.83,"AN": 21.0, "AO": 22.16,"AP": 25.33,"AQ": 16.83,
    "AR": 17.33,"AS": 18.16,"AT": 18.0, "AU": 18.16,"AV": 18.5, "AW": 15.83,
    "AX": 9.83,
}

LIGHT_BLUE = "DCE6F1"


# ─────────────────────────────────────────────
# PDF extraction
# ─────────────────────────────────────────────
def extract_pdf_text(pdf_path: str) -> str:
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if row:
                            parts.append(" | ".join(str(c) if c else "" for c in row))
            raw = page.extract_text()
            if raw:
                parts.append(raw)
    return "\n".join(parts)


# ─────────────────────────────────────────────
# Claude parsing
# ─────────────────────────────────────────────
SYSTEM_PROMPT = """Sei un esperto nell'estrarre dati da bolle di consegna fornitori italiani.
Dato il testo estratto da un PDF, individua TUTTI i prodotti/articoli presenti
e restituisci un array JSON strutturato. Non aggiungere testo al di fuori del JSON."""


def parse_with_claude(
    text: str,
    api_key: str,
    fornitore: str,
    produttore: str,
    iva: str,
    prezzi_ivati: bool,
) -> list[dict]:
    client = anthropic.Anthropic(api_key=api_key)

    nota_iva = (
        "I prezzi nel documento includono già l'IVA."
        if prezzi_ivati
        else f"I prezzi sono ESCLUSI IVA ({iva}%). Moltiplicali per {1 + int(iva) / 100:.2f} per ottenere il prezzo ivato."
    )

    user_prompt = f"""Analizza questo testo estratto da una bolla di consegna fornitore.
{nota_iva}

Restituisci SOLO un array JSON. Ogni elemento deve avere questi campi:
- "cod": codice articolo (stringa)
- "descrizione": descrizione prodotto (stringa)
- "ean": codice a barre EAN (stringa, null se assente)
- "quantita": quantità (intero)
- "prezzo_fornitore_ivato": prezzo di acquisto unitario CON IVA inclusa (float, 2 decimali)
- "prezzo_listino_ivato": prezzo di listino/vendita CON IVA inclusa (float o null)

Regole:
- Conserva gli zeri iniziali nei codici EAN/barcode
- Se ci sono sconti, applica lo sconto al prezzo prima di restituirlo
- Se il listino non è presente, usa null
- Quantità intera, prezzi float arrotondati a 2 decimali

Testo bolla:
{text}

Rispondi SOLO con l'array JSON."""

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=8192,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_prompt}],
    )

    raw = response.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


# ─────────────────────────────────────────────
# Excel generation — formato Danea esatto
# ─────────────────────────────────────────────
def create_danea_excel(
    products: list[dict],
    fornitore: str,
    produttore: str,
    iva: str,
    output_path: str,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foglio1"

    header_font = Font(name="Arial", bold=True, size=10)
    data_font   = Font(name="Arial", bold=False, size=10)
    blue_fill   = PatternFill("solid", fgColor=LIGHT_BLUE)
    no_fill     = PatternFill("solid", fgColor="FFFFFF")

    # Riga intestazione
    for ci, h in enumerate(DANEA_HEADERS, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Larghezze colonne
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Righe dati
    for ri, product in enumerate(products, start=2):
        fill = blue_fill if ri % 2 == 0 else no_fill  # righe pari = azzurro (come riferimento)

        row = [None] * 50
        cod = str(product.get("cod", "")).strip()

        row[0]  = cod                                              # A  Cod.
        row[1]  = str(product.get("descrizione", "")).strip()     # B  Descrizione
        row[6]  = iva                                              # G  Cod. Iva

        l1 = product.get("prezzo_listino_ivato")
        if l1 is not None:
            row[7] = round(float(l1), 2)                          # H  Listino 1 (ivato)

        ean = product.get("ean")
        if ean:
            row[14] = str(ean).strip()                            # O  Cod. a barre

        row[16] = produttore                                       # Q  Produttore
        row[25] = fornitore                                        # Z  Fornitore
        row[26] = cod                                              # AA Cod. prod. forn.

        pf = product.get("prezzo_fornitore_ivato")
        if pf is not None:
            row[27] = round(float(pf), 2)                         # AB Prezzo forn. (ivato)

        qty = int(product.get("quantita", 1))
        row[35] = qty                                              # AJ Q.tà giacenza
        row[37] = qty                                              # AL Q.tà disponibile

        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Altezza righe
    ws.row_dimensions[1].height = 15
    for ri in range(2, len(products) + 2):
        ws.row_dimensions[ri].height = 15

    wb.save(output_path)


# ─────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def frontend():
    html_path = Path(__file__).parent / "index.html"
    return html_path.read_text(encoding="utf-8")


@app.post("/convert")
async def convert_pdf(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    api_key: str = Form(...),
    fornitore: str = Form("ENESCO FRANCE"),
    produttore: str = Form("ENESCO"),
    cod_iva: str = Form("22"),
    prezzi_ivati: str = Form("false"),
):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Il file deve essere un PDF (.pdf)")

    is_ivato = prezzi_ivati.lower() == "true"
    file_id  = uuid.uuid4().hex
    pdf_path  = TEMP_DIR / f"{file_id}.pdf"
    xlsx_path = TEMP_DIR / f"{file_id}_danea.xlsx"

    try:
        pdf_path.write_bytes(await file.read())

        pdf_text = extract_pdf_text(str(pdf_path))
        if not pdf_text.strip():
            raise HTTPException(422, "Impossibile estrarre testo dal PDF. Prova un PDF non scansionato.")

        products = parse_with_claude(pdf_text, api_key, fornitore, produttore, cod_iva, is_ivato)
        if not products:
            raise HTTPException(422, "Nessun prodotto trovato nel PDF.")

        create_danea_excel(products, fornitore, produttore, cod_iva, str(xlsx_path))

        background_tasks.add_task(lambda: pdf_path.unlink(missing_ok=True))

        base_name = os.path.splitext(file.filename)[0]
        return JSONResponse({
            "success":  True,
            "file_id":  file_id,
            "prodotti": len(products),
            "filename": f"Danea_{base_name}.xlsx",
            "preview":  products[:5],   # anteprima prime 5 righe
        })

    except HTTPException:
        raise
    except anthropic.AuthenticationError:
        raise HTTPException(401, "API Key Anthropic non valida. Verifica la chiave e riprova.")
    except json.JSONDecodeError:
        raise HTTPException(422, "Errore nel parsing della risposta AI. Riprova.")
    except Exception as e:
        raise HTTPException(500, f"Errore: {str(e)}")


@app.get("/download/{file_id}")
async def download(file_id: str, background_tasks: BackgroundTasks, filename: str = "Danea_export"):
    xlsx_path = TEMP_DIR / f"{file_id}_danea.xlsx"
    if not xlsx_path.exists():
        raise HTTPException(404, "File non trovato o già scaricato.")

    background_tasks.add_task(lambda: xlsx_path.unlink(missing_ok=True))
    return FileResponse(
        path=str(xlsx_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{filename}.xlsx",
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
